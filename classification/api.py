"""
REST API for the classification system.
"""
from rest_framework import viewsets, status, permissions
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
from drf_spectacular.utils import extend_schema, OpenApiParameter

from .models import (
    TaxonomyCategory, Product, Classification,
    ClassificationAttribute, AlternativeCategory, ClassificationBatch
)
from .serializers import (
    TaxonomyCategorySerializer, ProductSerializer,
    ClassificationSerializer, ClassificationDetailSerializer,
    ClassificationAttributeSerializer, AlternativeCategorySerializer,
    ClassificationBatchSerializer, ClassificationStatsSerializer
)
# Lazy import - avoid loading ML libs at startup
def _get_classification_service():
    from .services import ClassificationService
    return ClassificationService


class TaxonomyCategoryViewSet(viewsets.ReadOnlyModelViewSet):
    """API endpoint for taxonomy categories."""
    
    queryset = TaxonomyCategory.objects.all()
    serializer_class = TaxonomyCategorySerializer
    permission_classes = [permissions.AllowAny]
    
    @extend_schema(
        summary="Get taxonomy tree",
        description="Retrieve the complete taxonomy hierarchy"
    )
    @action(detail=False, methods=['get'])
    def tree(self, request):
        """Get taxonomy as a tree structure."""
        root_categories = TaxonomyCategory.objects.filter(parent=None)
        tree = []
        
        for root in root_categories:
            tree.append(self._build_tree_node(root))
        
        return Response(tree)
    
    def _build_tree_node(self, category):
        """Build a tree node for a category."""
        children = category.children.all()
        return {
            'id': category.id,
            'shopify_id': category.shopify_id,
            'name': category.name,
            'full_path': category.full_path,
            'level': category.level,
            'children': [self._build_tree_node(child) for child in children]
        }


class ProductViewSet(viewsets.ModelViewSet):
    """API endpoint for products."""
    
    queryset = Product.objects.all()
    serializer_class = ProductSerializer
    permission_classes = [permissions.AllowAny]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status', 'classification_status']
    search_fields = ['product_name', 'product_number', 'source_category']
    
    @extend_schema(
        summary="Import products",
        description="Import products from Excel/CSV file"
    )
    @action(detail=False, methods=['post'])
    def import_products(self, request):
        """Import products from uploaded file."""
        if 'file' not in request.FILES:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file = request.FILES['file']
        
        # Process file based on extension
        if file.name.endswith('.xlsx') or file.name.endswith('.xls'):
            products = self._import_excel(file)
        elif file.name.endswith('.csv'):
            products = self._import_csv(file)
        else:
            return Response(
                {'error': 'Unsupported file format'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        return Response({
            'message': f'Imported {len(products)} products',
            'products': ProductSerializer(products, many=True).data
        })
    
    @extend_schema(
        summary="Export products",
        description="Export classified products to Excel"
    )
    @action(detail=False, methods=['get'])
    def export_products(self, request):
        """Export classified products."""
        # Placeholder - would generate Excel file in production
        return Response({'message': 'Export functionality'})
    
    def _import_excel(self, file):
        """Import products from Excel file."""
        import openpyxl
        
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        
        products = []
        headers = [cell.value for cell in ws[1]]
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            data = dict(zip(headers, row))
            
            product = Product.objects.create(
                product_number=data.get('Product Number', ''),
                product_name=data.get('Product Name', ''),
                product_description=data.get('Description', ''),
                image_url=data.get('Image URL', ''),
                source_category=data.get('Category', ''),
                materials=data.get('Materials', ''),
                product_weight=data.get('Weight'),
                country_of_origin=data.get('Country of Origin', '')
            )
            products.append(product)
        
        return products
    
    def _import_csv(self, file):
        """Import products from CSV file."""
        import csv
        
        products = []
        decoded_file = file.read().decode('utf-8')
        reader = csv.DictReader(decoded_file.splitlines())
        
        for row in reader:
            product = Product.objects.create(
                product_number=row.get('Product Number', ''),
                product_name=row.get('Product Name', ''),
                product_description=row.get('Description', ''),
                image_url=row.get('Image URL', ''),
                source_category=row.get('Category', ''),
                materials=row.get('Materials', ''),
                product_weight=row.get('Weight'),
                country_of_origin=row.get('Country of Origin', '')
            )
            products.append(product)
        
        return products


class ClassificationViewSet(viewsets.ModelViewSet):
    """API endpoint for classifications."""
    
    queryset = Classification.objects.all()
    permission_classes = [permissions.AllowAny]
    filter_backends = [DjangoFilterBackend]
    filterset_fields = ['status', 'requires_manual_review', 'classification_method']
    
    def get_serializer_class(self):
        if self.action == 'retrieve':
            return ClassificationDetailSerializer
        return ClassificationSerializer
    
    @extend_schema(
        summary="Classify product",
        description="Classify a single product"
    )
    @action(detail=False, methods=['post'])
    def classify(self, request):
        """Classify a single product."""
        product_id = request.data.get('product_id')
        use_llm = request.data.get('use_llm', False)
        
        try:
            product = Product.objects.get(id=product_id)
        except Product.DoesNotExist:
            return Response(
                {'error': 'Product not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get taxonomy categories
        categories = list(TaxonomyCategory.objects.values(
            'id', 'shopify_id', 'name', 'full_path',
            'keywords', 'product_type_hint', 'top_level_category'
        ))
        
        # Initialize classification service
        ClassificationService = _get_classification_service()
        service = ClassificationService(config={
            'taxonomy': categories,
            'llm_provider': 'groq',
            'llm_api_key': request.data.get('llm_api_key'),
        })
        service.initialize(categories)
        
        # Classify product
        product_data = {
            'id': product.id,
            'product_name': product.product_name,
            'product_description': product.product_description,
            'image_url': product.image_url,
            'source_category': product.source_category,
            'materials': product.materials,
            'product_weight': product.product_weight,
            'product_type': product.product_type
        }
        
        result = service.classify_product(product_data, use_llm=use_llm)
        
        return Response({
            'product_id': result.product_id,
            'category_path': result.category_path,
            'confidence': result.confidence,
            'confidence_level': result.confidence_level,
            'requires_review': result.requires_review,
            'review_reason': result.review_reason,
            'extracted_attributes': result.extracted_attributes,
            'alternatives': result.alternatives,
            'classification_method': result.classification_method,
            'reasoning': result.reasoning
        })
    
    @extend_schema(
        summary="Batch classify products",
        description="Classify multiple products in batch"
    )
    @action(detail=False, methods=['post'])
    def batch_classify(self, request):
        """Classify multiple products in batch."""
        product_ids = request.data.get('product_ids', [])
        use_llm = request.data.get('use_llm', False)
        
        products = Product.objects.filter(id__in=product_ids)
        
        # Get taxonomy categories
        categories = list(TaxonomyCategory.objects.values(
            'id', 'shopify_id', 'name', 'full_path',
            'keywords', 'product_type_hint', 'top_level_category'
        ))
        
        # Initialize classification service
        ClassificationService = _get_classification_service()
        service = ClassificationService(config={
            'taxonomy': categories,
            'llm_provider': 'groq',
            'llm_api_key': request.data.get('llm_api_key'),
        })
        service.initialize(categories)
        
        # Classify products
        results = []
        for product in products:
            product_data = {
                'id': product.id,
                'product_name': product.product_name,
                'product_description': product.product_description,
                'image_url': product.image_url,
                'source_category': product.source_category,
                'materials': product.materials,
                'product_weight': product.product_weight,
                'product_type': product.product_type
            }
            
            result = service.classify_product(product_data, use_llm=use_llm)
            results.append({
                'product_id': result.product_id,
                'category_path': result.category_path,
                'confidence': result.confidence,
                'requires_review': result.requires_review
            })
        
        return Response({
            'total': len(results),
            'results': results
        })
    
    @extend_schema(
        summary="Approve classification",
        description="Approve a classification"
    )
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """Approve a classification."""
        classification = self.get_object()
        reviewed_by = request.data.get('reviewed_by', 'system')
        notes = request.data.get('notes', '')
        
        classification.approve(reviewed_by=reviewed_by, notes=notes)
        
        return Response({'message': 'Classification approved'})
    
    @extend_schema(
        summary="Reject classification",
        description="Reject a classification"
    )
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """Reject a classification."""
        classification = self.get_object()
        reviewed_by = request.data.get('reviewed_by', 'system')
        notes = request.data.get('notes', '')
        
        classification.reject(reviewed_by=reviewed_by, notes=notes)
        
        return Response({'message': 'Classification rejected'})
    
    @extend_schema(
        summary="Reassign classification",
        description="Reassign product to different category"
    )
    @action(detail=True, methods=['post'])
    def reassign(self, request, pk=None):
        """Reassign product to different category."""
        classification = self.get_object()
        new_category_path = request.data.get('category_path')
        
        if not new_category_path:
            return Response(
                {'error': 'Category path required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Find category
        try:
            new_category = TaxonomyCategory.objects.get(full_path=new_category_path)
        except TaxonomyCategory.DoesNotExist:
            return Response(
                {'error': 'Category not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Update classification
        classification.taxonomy_node = new_category
        classification.predicted_category_path = new_category_path
        classification.status = 'approved'
        classification.reviewed_by = request.data.get('reviewed_by', 'system')
        classification.reviewed_at = timezone.now()
        classification.save()
        
        return Response({'message': 'Classification reassigned'})
    
    @extend_schema(
        summary="Batch approve",
        description="Approve multiple classifications"
    )
    @action(detail=False, methods=['post'])
    def batch_approve(self, request):
        """Approve multiple classifications."""
        classification_ids = request.data.get('classification_ids', [])
        reviewed_by = request.data.get('reviewed_by', 'system')
        
        classifications = Classification.objects.filter(id__in=classification_ids)
        count = classifications.count()
        
        classifications.update(
            status='approved',
            reviewed_by=reviewed_by,
            reviewed_at=timezone.now(),
            requires_manual_review=False
        )
        
        return Response({
            'message': f'Approved {count} classifications'
        })


class ClassificationBatchViewSet(viewsets.ModelViewSet):
    """API endpoint for classification batches."""
    
    queryset = ClassificationBatch.objects.all()
    serializer_class = ClassificationBatchSerializer
    permission_classes = [permissions.AllowAny]
    
    @extend_schema(
        summary="Start batch",
        description="Start batch classification"
    )
    @action(detail=True, methods=['post'])
    def start(self, request, pk=None):
        """Start batch classification."""
        batch = self.get_object()
        batch.start()
        
        return Response({'message': 'Batch started'})
    
    @extend_schema(
        summary="Get batch progress",
        description="Get batch classification progress"
    )
    @action(detail=True, methods=['get'])
    def progress(self, request, pk=None):
        """Get batch progress."""
        batch = self.get_object()
        
        return Response({
            'status': batch.status,
            'total_products': batch.total_products,
            'processed_products': batch.processed_products,
            'progress_percentage': batch.progress_percentage,
            'successful_classifications': batch.successful_classifications,
            'failed_classifications': batch.failed_classifications,
            'reviews_needed': batch.reviews_needed
        })


class StatsViewSet(viewsets.ViewSet):
    """API endpoint for statistics."""
    
    permission_classes = [permissions.AllowAny]
    
    @extend_schema(
        summary="Get classification stats",
        description="Get overall classification statistics"
    )
    def list(self, request):
        """Get classification statistics."""
        total_products = Product.objects.count()
        classified = Classification.objects.exclude(status='pending').count()
        needs_review = Classification.objects.filter(requires_manual_review=True).count()
        approved = Classification.objects.filter(status='approved').count()
        
        avg_confidence = 0.0
        if classified > 0:
            from django.db.models import Avg
            avg_confidence = Classification.objects.aggregate(
                avg_confidence=Avg('confidence')
            )['avg_confidence'] or 0.0
        
        status_counts = {}
        for status, _ in Classification.STATUS_CHOICES:
            status_counts[status] = Classification.objects.filter(status=status).count()
        
        return Response({
            'total_products': total_products,
            'classified': classified,
            'needs_review': needs_review,
            'approved': approved,
            'average_confidence': round(avg_confidence, 3),
            'status_counts': status_counts
        })
